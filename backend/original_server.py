from fastapi import (
    FastAPI,
    APIRouter,
    Depends,
    HTTPException,
    status,
    Request,
    File,
    UploadFile,
    Form,
)
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from fastapi.staticfiles import StaticFiles
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
import os
import logging
import uuid
import bcrypt
import jwt
import pyotp
import qrcode
import io
import base64
import json
import shutil
from datetime import datetime, timedelta, timezone, date
from typing import List, Optional, Dict, Any, Union
from pydantic import BaseModel, Field, EmailStr
from enum import Enum
from fastapi.responses import StreamingResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

# Import inventory models and services
from inventory_models import *
from inventory_service import InventoryService

# Import financial models and services
from financial_models import *
from financial_service import FinancialService

# Import notification and automation services
from notification_models import *
from notification_service import NotificationService
from automation_service import AutomationService

# Import billing system
from billing_models import *
from billing_service import BillingService
from stripe_integration import StripeIntegrationService


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# Create uploads directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# Security configuration
SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7

# FastAPI app setup
app = FastAPI(title="COAM SaaS API", version="1.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# Initialize services
inventory_service = None
financial_service = None
notification_service = None
automation_service = None


# Enums
class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    ML_ADMIN = "ml_admin"  # Master Licensee Admin
    DISPATCH = "dispatch"
    TECH = "tech"
    WAREHOUSE = "warehouse"
    ACCOUNTANT = "accountant"


class JobType(str, Enum):
    INSTALLATION = "installation"
    MAINTENANCE = "maintenance"
    REPAIR = "repair"
    INSPECTION = "inspection"
    REMOVAL = "removal"
    TRANSFER = "transfer"


class JobStatus(str, Enum):
    PENDING = "pending"
    ASSIGNED = "assigned"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    CANCELLED = "cancelled"
    ON_HOLD = "on_hold"


class JobPriority(str, Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    URGENT = "urgent"


class MachineStatus(str, Enum):
    ACTIVE = "active"
    INACTIVE = "inactive"
    MAINTENANCE = "maintenance"
    BROKEN = "broken"
    TRANSFERRED = "transferred"


class PartStatus(str, Enum):
    AVAILABLE = "available"
    RESERVED = "reserved"
    IN_USE = "in_use"
    OUT_OF_STOCK = "out_of_stock"


class NotificationType(str, Enum):
    JOB_ASSIGNED = "job_assigned"
    JOB_APPROVED = "job_approved"
    URGENT_JOB = "urgent_job"
    LOW_INVENTORY = "low_inventory"
    PART_REPLENISHMENT = "part_replenishment"
    RMA_UPDATE = "rma_update"
    SYSTEM_ALERT = "system_alert"


class AuditAction(str, Enum):
    CREATE = "create"
    UPDATE = "update"
    DELETE = "delete"
    LOGIN = "login"
    LOGOUT = "logout"
    FAILED_LOGIN = "failed_login"
    CHECK_IN = "check_in"
    CHECK_OUT = "check_out"


# Pydantic Models
class Tenant(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company_name: str
    contact_email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    license_number: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True
    settings: Dict[str, Any] = Field(default_factory=dict)


class TechProfile(BaseModel):
    certifications: List[str] = Field(default_factory=list)
    specializations: List[str] = Field(default_factory=list)
    experience_level: str = "junior"  # junior, intermediate, senior
    restrictions: List[str] = Field(default_factory=list)
    max_jobs_per_day: int = 8
    preferred_regions: List[str] = Field(default_factory=list)


class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: Optional[str] = None  # None for super_admin
    email: EmailStr
    first_name: str
    last_name: str
    role: UserRole
    phone: Optional[str] = None
    is_active: bool = True
    mfa_enabled: bool = False
    mfa_secret: Optional[str] = None
    tech_profile: Optional[TechProfile] = None
    current_location: Optional[Dict[str, float]] = None  # {"lat": 0.0, "lng": 0.0}
    last_check_in: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_login: Optional[datetime] = None


class UserCreate(BaseModel):
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    role: UserRole
    phone: Optional[str] = None
    tenant_id: Optional[str] = None
    tech_profile: Optional[TechProfile] = None


class UserLogin(BaseModel):
    email: EmailStr
    password: str
    mfa_code: Optional[str] = None


class Location(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    name: str
    address: str
    city: str
    state: str = "Georgia"
    zip_code: str
    contact_person: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    coordinates: Optional[Dict[str, float]] = None  # {"lat": 0.0, "lng": 0.0}
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Machine(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    serial_number: str
    model: str
    manufacturer: str
    location_id: str
    status: MachineStatus = MachineStatus.ACTIVE
    qr_code: Optional[str] = None
    barcode: Optional[str] = None
    installed_date: Optional[datetime] = None
    last_maintenance: Optional[datetime] = None
    next_maintenance: Optional[datetime] = None
    service_history: List[Dict[str, Any]] = Field(default_factory=list)
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Part(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    part_number: str
    name: str
    description: Optional[str] = None
    category: str
    unit_cost: float
    barcode: Optional[str] = None
    qr_code: Optional[str] = None
    supplier: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PartUsage(BaseModel):
    part_id: str
    quantity_used: int
    unit_cost: float
    notes: Optional[str] = None


class TruckInventory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    tech_id: str
    part_id: str
    quantity: int
    reserved_quantity: int = 0
    last_updated: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Inventory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    part_id: str
    quantity_on_hand: int
    minimum_quantity: int = 0
    reserved_quantity: int = 0
    location: str = "warehouse"
    status: PartStatus = PartStatus.AVAILABLE
    last_updated: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class JobCheckIn(BaseModel):
    location: Dict[str, float]  # {"lat": 0.0, "lng": 0.0}
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = None


class JobFile(BaseModel):
    file_id: str
    file_type: str  # "photo", "video", "invoice", "document"
    file_path: str
    description: Optional[str] = None
    ocr_data: Optional[Dict[str, Any]] = None
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Job(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    job_type: JobType
    status: JobStatus = JobStatus.PENDING
    priority: JobPriority = JobPriority.MEDIUM
    machine_id: Optional[str] = None
    location_id: str
    assigned_tech_id: Optional[str] = None
    created_by_id: str
    approved_by_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    required_skills: List[str] = Field(default_factory=list)
    estimated_duration: Optional[int] = None  # minutes
    scheduled_date: Optional[datetime] = None
    completed_date: Optional[datetime] = None
    check_in: Optional[JobCheckIn] = None
    check_out: Optional[JobCheckIn] = None
    parts_required: List[str] = Field(default_factory=list)  # part_ids
    parts_used: List[PartUsage] = Field(default_factory=list)
    files: List[JobFile] = Field(default_factory=list)
    invoice_total: Optional[float] = None
    tech_notes: Optional[str] = None
    dispatch_notes: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Contract(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    location_id: str
    contract_number: str
    start_date: datetime
    end_date: datetime
    revenue_split: float  # Percentage for ML
    terms: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: Optional[str] = None
    user_id: str
    notification_type: NotificationType
    title: str
    message: str
    data: Optional[Dict[str, Any]] = None
    read: bool = False
    sent_via_email: bool = False
    sent_via_sms: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class AuditLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: Optional[str] = None
    user_id: str
    action: AuditAction
    resource_type: str
    resource_id: Optional[str] = None
    old_values: Optional[Dict[str, Any]] = None
    new_values: Optional[Dict[str, Any]] = None
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    location: Optional[Dict[str, float]] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# Request/Response Models
class JobAssignment(BaseModel):
    job_ids: List[str]
    tech_id: str
    scheduled_date: Optional[datetime] = None
    notes: Optional[str] = None


class BulkJobApproval(BaseModel):
    assignments: List[JobAssignment]


class LocationCheckIn(BaseModel):
    job_id: str
    location: Dict[str, float]
    notes: Optional[str] = None


class PartCheckout(BaseModel):
    job_id: str
    parts: List[Dict[str, Union[str, int]]]  # [{"part_id": "...", "quantity": 2}]


class OCRResult(BaseModel):
    text: str
    confidence: float
    structured_data: Dict[str, Any]


# Authentication utilities
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(
            minutes=ACCESS_TOKEN_EXPIRE_MINUTES
        )
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def create_refresh_token(user_id: str):
    expire = datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode = {"sub": user_id, "exp": expire, "type": "refresh"}
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")

        user_data = await db.users.find_one({"id": user_id})
        if user_data is None:
            raise HTTPException(status_code=401, detail="User not found")

        return User(**user_data)
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


# Multi-tenant middleware
class TenantMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        # Extract tenant context from user if authenticated
        if hasattr(request.state, "user"):
            request.state.tenant_id = getattr(request.state.user, "tenant_id", None)
        else:
            request.state.tenant_id = None

        response = await call_next(request)
        return response


# Utility functions
async def log_audit(
    user_id: str,
    action: AuditAction,
    resource_type: str,
    resource_id: Optional[str] = None,
    old_values: Optional[Dict] = None,
    new_values: Optional[Dict] = None,
    tenant_id: Optional[str] = None,
    ip_address: Optional[str] = None,
    location: Optional[Dict[str, float]] = None,
):
    audit_log = AuditLog(
        tenant_id=tenant_id,
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        old_values=old_values,
        new_values=new_values,
        ip_address=ip_address,
        location=location,
    )
    await db.audit_logs.insert_one(audit_log.dict())


async def send_notification(
    user_id: str,
    notification_type: NotificationType,
    title: str,
    message: str,
    data: Optional[Dict] = None,
    tenant_id: Optional[str] = None,
):
    notification = Notification(
        tenant_id=tenant_id,
        user_id=user_id,
        notification_type=notification_type,
        title=title,
        message=message,
        data=data or {},
    )
    await db.notifications.insert_one(notification.dict())
    return notification


def check_permission(
    user: User, required_role: UserRole, tenant_id: Optional[str] = None
):
    # Super admin has access to everything
    if user.role == UserRole.SUPER_ADMIN:
        return True

    # Check tenant access for regular users
    if tenant_id and user.tenant_id != tenant_id:
        return False

    # Role hierarchy check
    role_hierarchy = {
        UserRole.SUPER_ADMIN: 6,
        UserRole.ML_ADMIN: 5,
        UserRole.DISPATCH: 4,
        UserRole.TECH: 3,
        UserRole.WAREHOUSE: 2,
        UserRole.ACCOUNTANT: 1,
    }

    return role_hierarchy.get(user.role, 0) >= role_hierarchy.get(required_role, 0)


# Authentication endpoints
@api_router.post("/auth/register")
async def register_user(
    user_data: UserCreate, current_user: User = Depends(get_current_user)
):
    # Only ML_ADMIN and SUPER_ADMIN can create users
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")

    # Set tenant_id based on current user
    if current_user.role != UserRole.SUPER_ADMIN:
        user_data.tenant_id = current_user.tenant_id

    # Hash password and create user
    user_dict = user_data.dict(exclude={"password"})
    user = User(**user_dict)

    # Add password hash to the document that will be stored
    user_doc = user.dict()
    user_doc["password_hash"] = hash_password(user_data.password)

    await db.users.insert_one(user_doc)

    # Log audit
    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "user",
        user.id,
        tenant_id=current_user.tenant_id,
    )

    return {"message": "User created successfully", "user_id": user.id}


@api_router.post("/auth/login")
async def login(user_credentials: UserLogin, request: Request):
    user_data = await db.users.find_one({"email": user_credentials.email})
    if not user_data:
        await log_audit(
            "unknown", AuditAction.FAILED_LOGIN, "auth", ip_address=request.client.host
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")

    user = User(**user_data)

    # Verify password
    if not verify_password(user_credentials.password, user_data["password_hash"]):
        await log_audit(
            user.id,
            AuditAction.FAILED_LOGIN,
            "auth",
            tenant_id=user.tenant_id,
            ip_address=request.client.host,
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Check MFA if enabled
    if user.mfa_enabled:
        if not user_credentials.mfa_code:
            raise HTTPException(status_code=400, detail="MFA code required")

        totp = pyotp.TOTP(user.mfa_secret)
        if not totp.verify(user_credentials.mfa_code):
            await log_audit(
                user.id,
                AuditAction.FAILED_LOGIN,
                "auth",
                tenant_id=user.tenant_id,
                ip_address=request.client.host,
            )
            raise HTTPException(status_code=401, detail="Invalid MFA code")

    # Create tokens
    access_token = create_access_token(data={"sub": user.id})
    refresh_token = create_refresh_token(user.id)

    # Update last login
    await db.users.update_one(
        {"id": user.id}, {"$set": {"last_login": datetime.now(timezone.utc)}}
    )

    # Log successful login
    await log_audit(
        user.id,
        AuditAction.LOGIN,
        "auth",
        tenant_id=user.tenant_id,
        ip_address=request.client.host,
    )

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": user.dict(),
    }


@api_router.post("/auth/setup-mfa")
async def setup_mfa(current_user: User = Depends(get_current_user)):
    if current_user.mfa_enabled:
        raise HTTPException(status_code=400, detail="MFA already enabled")

    # Generate MFA secret
    secret = pyotp.random_base32()

    # Generate QR code
    totp_uri = pyotp.totp.TOTP(secret).provisioning_uri(
        name=current_user.email, issuer_name="COAM SaaS"
    )

    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(totp_uri)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_code_data = base64.b64encode(buffer.getvalue()).decode()

    # Save secret (temporarily)
    await db.users.update_one({"id": current_user.id}, {"$set": {"mfa_secret": secret}})

    return {"secret": secret, "qr_code": f"data:image/png;base64,{qr_code_data}"}


# Job Management Endpoints
@api_router.post("/jobs", response_model=Job)
async def create_job(job_data: dict, current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    job_data["tenant_id"] = current_user.tenant_id
    job_data["created_by_id"] = current_user.id

    job = Job(**job_data)
    await db.jobs.insert_one(job.dict())

    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "job",
        job.id,
        tenant_id=current_user.tenant_id,
    )

    return job


@api_router.get("/jobs", response_model=List[Job])
async def get_jobs(
    status: Optional[JobStatus] = None,
    priority: Optional[JobPriority] = None,
    assigned_tech_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    # Filter for techs to see only their assigned jobs
    if current_user.role == UserRole.TECH:
        query["assigned_tech_id"] = current_user.id

    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if assigned_tech_id:
        query["assigned_tech_id"] = assigned_tech_id

    jobs = await db.jobs.find(query).sort("created_at", -1).to_list(1000)
    return [Job(**job) for job in jobs]


@api_router.get("/jobs/{job_id}", response_model=Job)
async def get_job(job_id: str, current_user: User = Depends(get_current_user)):
    query = {"id": job_id}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    job_data = await db.jobs.find_one(query)
    if not job_data:
        raise HTTPException(status_code=404, detail="Job not found")

    return Job(**job_data)


@api_router.put("/jobs/{job_id}/assign")
async def assign_job(
    job_id: str,
    assignment: JobAssignment,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Update job assignment
    update_data = {
        "assigned_tech_id": assignment.tech_id,
        "status": JobStatus.ASSIGNED,
        "updated_at": datetime.now(timezone.utc),
    }

    if assignment.scheduled_date:
        update_data["scheduled_date"] = assignment.scheduled_date
    if assignment.notes:
        update_data["dispatch_notes"] = assignment.notes

    result = await db.jobs.update_one(
        {"id": job_id, "tenant_id": current_user.tenant_id}, {"$set": update_data}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")

    # Send notification to tech
    await send_notification(
        assignment.tech_id,
        NotificationType.JOB_ASSIGNED,
        "New Job Assigned",
        f"You have been assigned a new job: {job_id}",
        {"job_id": job_id},
        current_user.tenant_id,
    )

    await log_audit(
        current_user.id,
        AuditAction.UPDATE,
        "job",
        job_id,
        tenant_id=current_user.tenant_id,
    )

    return {"message": "Job assigned successfully"}


@api_router.post("/jobs/bulk-approve")
async def bulk_approve_jobs(
    approval: BulkJobApproval, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    results = []
    for assignment in approval.assignments:
        for job_id in assignment.job_ids:
            # Check parts availability and tech restrictions
            job_data = await db.jobs.find_one(
                {"id": job_id, "tenant_id": current_user.tenant_id}
            )
            if not job_data:
                continue

            job = Job(**job_data)

            # Check tech skills vs required skills
            tech_data = await db.users.find_one(
                {"id": assignment.tech_id, "tenant_id": current_user.tenant_id}
            )
            if tech_data and job.required_skills:
                tech = User(**tech_data)
                if tech.tech_profile:
                    tech_skills = (
                        tech.tech_profile.specializations
                        + tech.tech_profile.certifications
                    )
                    missing_skills = [
                        skill
                        for skill in job.required_skills
                        if skill not in tech_skills
                    ]
                    if missing_skills:
                        results.append(
                            {
                                "job_id": job_id,
                                "status": "warning",
                                "message": f"Tech missing skills: {', '.join(missing_skills)}",
                            }
                        )

            # Update job
            update_data = {
                "assigned_tech_id": assignment.tech_id,
                "status": JobStatus.ASSIGNED,
                "approved_by_id": current_user.id,
                "updated_at": datetime.now(timezone.utc),
            }

            if assignment.scheduled_date:
                update_data["scheduled_date"] = assignment.scheduled_date

            await db.jobs.update_one(
                {"id": job_id, "tenant_id": current_user.tenant_id},
                {"$set": update_data},
            )

            # Send notification
            await send_notification(
                assignment.tech_id,
                NotificationType.JOB_APPROVED,
                "Job Approved & Assigned",
                f"Job {job_id} has been approved and assigned to you",
                {"job_id": job_id},
                current_user.tenant_id,
            )

            results.append(
                {
                    "job_id": job_id,
                    "status": "success",
                    "message": "Assigned successfully",
                }
            )

    return {"results": results}


# Tech Mobile App Endpoints
@api_router.post("/tech/checkin")
async def tech_check_in(
    checkin: LocationCheckIn, current_user: User = Depends(get_current_user)
):
    if current_user.role != UserRole.TECH:
        raise HTTPException(status_code=403, detail="Only techs can check in")

    # Update job with check-in info
    check_in_data = JobCheckIn(location=checkin.location, notes=checkin.notes)

    result = await db.jobs.update_one(
        {"id": checkin.job_id, "assigned_tech_id": current_user.id},
        {
            "$set": {
                "check_in": check_in_data.dict(),
                "status": JobStatus.IN_PROGRESS,
                "updated_at": datetime.now(timezone.utc),
            }
        },
    )

    if result.matched_count == 0:
        raise HTTPException(
            status_code=404, detail="Job not found or not assigned to you"
        )

    # Update user location
    await db.users.update_one(
        {"id": current_user.id},
        {
            "$set": {
                "current_location": checkin.location,
                "last_check_in": datetime.now(timezone.utc),
            }
        },
    )

    await log_audit(
        current_user.id,
        AuditAction.CHECK_IN,
        "job",
        checkin.job_id,
        tenant_id=current_user.tenant_id,
        location=checkin.location,
    )

    return {"message": "Checked in successfully"}


@api_router.post("/tech/checkout")
async def tech_check_out(
    checkout: LocationCheckIn, current_user: User = Depends(get_current_user)
):
    if current_user.role != UserRole.TECH:
        raise HTTPException(status_code=403, detail="Only techs can check out")

    check_out_data = JobCheckIn(location=checkout.location, notes=checkout.notes)

    result = await db.jobs.update_one(
        {"id": checkout.job_id, "assigned_tech_id": current_user.id},
        {
            "$set": {
                "check_out": check_out_data.dict(),
                "status": JobStatus.COMPLETED,
                "completed_date": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            }
        },
    )

    if result.matched_count == 0:
        raise HTTPException(
            status_code=404, detail="Job not found or not assigned to you"
        )

    await log_audit(
        current_user.id,
        AuditAction.CHECK_OUT,
        "job",
        checkout.job_id,
        tenant_id=current_user.tenant_id,
        location=checkout.location,
    )

    return {"message": "Checked out successfully"}


@api_router.post("/tech/parts/checkout")
async def checkout_parts(
    checkout: PartCheckout, current_user: User = Depends(get_current_user)
):
    if current_user.role != UserRole.TECH:
        raise HTTPException(status_code=403, detail="Only techs can checkout parts")

    # Update truck inventory
    for part in checkout.parts:
        await db.truck_inventory.update_one(
            {"tech_id": current_user.id, "part_id": part["part_id"]},
            {
                "$inc": {
                    "quantity": -part["quantity"],
                    "reserved_quantity": part["quantity"],
                }
            },
            upsert=False,
        )

    # Update job with parts used
    parts_used = [
        PartUsage(
            part_id=part["part_id"],
            quantity_used=part["quantity"],
            unit_cost=0.0,  # Will be updated when cost is known
            notes="",
        )
        for part in checkout.parts
    ]

    await db.jobs.update_one(
        {"id": checkout.job_id, "assigned_tech_id": current_user.id},
        {"$push": {"parts_used": {"$each": [part.dict() for part in parts_used]}}},
    )

    return {"message": "Parts checked out successfully"}


@api_router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    job_id: str = Form(...),
    file_type: str = Form(...),
    description: str = Form(None),
    current_user: User = Depends(get_current_user),
):
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "video/mp4", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File type not supported")

    # Save file
    file_id = str(uuid.uuid4())
    file_extension = file.filename.split(".")[-1]
    file_path = UPLOAD_DIR / f"{file_id}.{file_extension}"

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Create job file record
    job_file = JobFile(
        file_id=file_id,
        file_type=file_type,
        file_path=str(file_path.relative_to(ROOT_DIR)),
        description=description,
    )

    # Update job with file
    await db.jobs.update_one(
        {"id": job_id, "tenant_id": current_user.tenant_id},
        {"$push": {"files": job_file.dict()}},
    )

    return {"file_id": file_id, "message": "File uploaded successfully"}


# Notifications
@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(current_user: User = Depends(get_current_user)):
    notifications = (
        await db.notifications.find({"user_id": current_user.id})
        .sort("created_at", -1)
        .limit(50)
        .to_list(length=None)
    )

    return [Notification(**notif) for notif in notifications]


@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str, current_user: User = Depends(get_current_user)
):
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user.id}, {"$set": {"read": True}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")

    return {"message": "Notification marked as read"}


# Inventory Management
@api_router.get("/inventory", response_model=List[Inventory])
async def get_inventory(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    inventory = await db.inventory.find(query).to_list(1000)
    return [Inventory(**item) for item in inventory]


@api_router.get("/tech/truck-inventory")
async def get_truck_inventory(current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.TECH:
        raise HTTPException(
            status_code=403, detail="Only techs can view truck inventory"
        )

    truck_inventory = await db.truck_inventory.find(
        {"tech_id": current_user.id, "tenant_id": current_user.tenant_id}
    ).to_list(1000)

    return [TruckInventory(**item) for item in truck_inventory]


# Parts Management
@api_router.post("/parts", response_model=Part)
async def create_part(part_data: dict, current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    part_data["tenant_id"] = current_user.tenant_id
    part = Part(**part_data)
    await db.parts.insert_one(part.dict())

    return part


@api_router.get("/parts", response_model=List[Part])
async def get_parts(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    parts = await db.parts.find(query).to_list(1000)
    return [Part(**part) for part in parts]


# Machine Management (Extended)
@api_router.post("/machines", response_model=Machine)
async def create_machine(
    machine_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    machine_data["tenant_id"] = current_user.tenant_id

    # Generate QR code and barcode if not provided
    if not machine_data.get("qr_code"):
        machine_data["qr_code"] = f"COAM-{machine_data['serial_number']}"
    if not machine_data.get("barcode"):
        machine_data["barcode"] = machine_data["serial_number"]

    machine = Machine(**machine_data)
    await db.machines.insert_one(machine.dict())

    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "machine",
        machine.id,
        tenant_id=current_user.tenant_id,
    )

    return machine


@api_router.get("/machines", response_model=List[Machine])
async def get_machines(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    machines = await db.machines.find(query).to_list(1000)
    return [Machine(**machine) for machine in machines]


@api_router.get("/machines/scan/{code}")
async def scan_machine(code: str, current_user: User = Depends(get_current_user)):
    query = {"$or": [{"qr_code": code}, {"barcode": code}, {"serial_number": code}]}

    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    machine_data = await db.machines.find_one(query)
    if not machine_data:
        raise HTTPException(status_code=404, detail="Machine not found")

    return Machine(**machine_data)


# Location Management (Extended)
@api_router.post("/locations", response_model=Location)
async def create_location(
    location_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    location_data["tenant_id"] = current_user.tenant_id
    location = Location(**location_data)
    await db.locations.insert_one(location.dict())

    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "location",
        location.id,
        tenant_id=current_user.tenant_id,
    )

    return location


@api_router.get("/locations", response_model=List[Location])
async def get_locations(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    locations = await db.locations.find(query).to_list(1000)
    return [Location(**location) for location in locations]


# Tenant management
@api_router.post("/tenants", response_model=Tenant)
async def create_tenant(
    tenant_data: dict, current_user: User = Depends(get_current_user)
):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(
            status_code=403, detail="Only super admin can create tenants"
        )

    tenant = Tenant(**tenant_data)
    await db.tenants.insert_one(tenant.dict())

    await log_audit(current_user.id, AuditAction.CREATE, "tenant", tenant.id)

    return tenant


@api_router.get("/tenants", response_model=List[Tenant])
async def get_tenants(current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    tenants = await db.tenants.find().to_list(1000)
    return [Tenant(**tenant) for tenant in tenants]


# Admin dashboard endpoints
@api_router.get("/admin/dashboard")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    if current_user.role == UserRole.SUPER_ADMIN:
        # Super admin sees system-wide stats
        stats = {
            "total_tenants": await db.tenants.count_documents({}),
            "total_users": await db.users.count_documents({}),
            "total_machines": await db.machines.count_documents({}),
            "total_locations": await db.locations.count_documents({}),
            "recent_logins": await db.audit_logs.count_documents(
                {
                    "action": AuditAction.LOGIN,
                    "timestamp": {
                        "$gte": datetime.now(timezone.utc) - timedelta(days=1)
                    },
                }
            ),
        }
    else:
        # Tenant admin sees tenant-specific stats
        if not check_permission(current_user, UserRole.ML_ADMIN):
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        query = {"tenant_id": current_user.tenant_id}
        stats = {
            "total_users": await db.users.count_documents(query),
            "total_machines": await db.machines.count_documents(query),
            "total_locations": await db.locations.count_documents(query),
            "total_jobs": await db.jobs.count_documents(query),
            "pending_jobs": await db.jobs.count_documents(
                {**query, "status": JobStatus.PENDING}
            ),
            "in_progress_jobs": await db.jobs.count_documents(
                {**query, "status": JobStatus.IN_PROGRESS}
            ),
            "active_machines": await db.machines.count_documents(
                {**query, "status": MachineStatus.ACTIVE}
            ),
            "active_techs": await db.users.count_documents(
                {**query, "role": UserRole.TECH, "is_active": True}
            ),
        }

    return stats


# Enhanced Asset Management Endpoints
@api_router.post("/assets", response_model=EnhancedAsset)
async def create_asset(
    asset_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Generate asset tag
    asset_type = AssetType(asset_data.get("asset_type"))
    asset_tag = await inventory_service.generate_asset_tag(
        asset_type, current_user.tenant_id
    )

    asset_data["tenant_id"] = current_user.tenant_id
    asset_data["asset_tag"] = asset_tag.dict()

    asset = EnhancedAsset(**asset_data)
    await db.assets.insert_one(asset.dict())

    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "asset",
        asset.id,
        tenant_id=current_user.tenant_id,
    )

    return asset


@api_router.get("/assets", response_model=List[EnhancedAsset])
async def get_assets(
    asset_type: Optional[AssetType] = None,
    status: Optional[AssetStatus] = None,
    current_user: User = Depends(get_current_user),
):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    if asset_type:
        query["asset_type"] = asset_type.value
    if status:
        query["status"] = status.value

    assets = await db.assets.find(query).to_list(1000)
    return [EnhancedAsset(**asset) for asset in assets]


@api_router.get(
    "/assets/{asset_id}/lifecycle", response_model=MachineLifecycleAnalytics
)
async def get_asset_lifecycle(
    asset_id: str, current_user: User = Depends(get_current_user)
):
    if current_user.role == UserRole.TECH:
        # Techs can only view assets they're working on
        job = await db.jobs.find_one(
            {
                "assigned_tech_id": current_user.id,
                "machine_id": asset_id,
                "status": {"$in": ["assigned", "in_progress"]},
            }
        )
        if not job:
            raise HTTPException(
                status_code=403, detail="Not authorized to view this asset"
            )

    try:
        analytics = await inventory_service.track_machine_lifecycle(asset_id)
        return analytics
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@api_router.post("/assets/{asset_id}/maintenance")
async def schedule_asset_maintenance(
    asset_id: str,
    maintenance_request: AssetMaintenanceRequest,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Create lifecycle event
    lifecycle_event = AssetLifecycle(
        asset_id=asset_id,
        event_type="maintenance_scheduled",
        performed_by=current_user.id,
        notes=maintenance_request.notes,
        next_service_date=maintenance_request.scheduled_date,
    )

    await db.asset_lifecycle.insert_one(lifecycle_event.dict())

    # Update asset next maintenance date
    await db.assets.update_one(
        {"id": asset_id, "tenant_id": current_user.tenant_id},
        {"$set": {"next_maintenance": maintenance_request.scheduled_date}},
    )

    return {
        "message": "Maintenance scheduled successfully",
        "event_id": lifecycle_event.id,
    }


# Enhanced Inventory Management
@api_router.get("/inventory/dashboard", response_model=InventoryDashboardStats)
async def get_inventory_dashboard(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    stats = await inventory_service.calculate_dashboard_stats(current_user.tenant_id)
    return stats


@api_router.post("/inventory/transfer")
async def transfer_inventory(
    transfer_request: InventoryTransferRequest,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Get source inventory
    source_inventory = await db.inventory.find_one(
        {
            "tenant_id": current_user.tenant_id,
            "part_id": transfer_request.part_id,
            "location": transfer_request.from_location.value,
        }
    )

    if (
        not source_inventory
        or source_inventory["quantity_on_hand"] < transfer_request.quantity
    ):
        raise HTTPException(
            status_code=400, detail="Insufficient inventory for transfer"
        )

    # Deduct from source
    await db.inventory.update_one(
        {"id": source_inventory["id"]},
        {
            "$inc": {"quantity_on_hand": -transfer_request.quantity},
            "$set": {"updated_at": datetime.now(timezone.utc)},
        },
    )

    # Add to destination
    dest_inventory = await db.inventory.find_one(
        {
            "tenant_id": current_user.tenant_id,
            "part_id": transfer_request.part_id,
            "location": transfer_request.to_location.value,
        }
    )

    if dest_inventory:
        await db.inventory.update_one(
            {"id": dest_inventory["id"]},
            {
                "$inc": {"quantity_on_hand": transfer_request.quantity},
                "$set": {"updated_at": datetime.now(timezone.utc)},
            },
        )
    else:
        # Create new inventory record at destination
        new_inventory = EnhancedInventory(
            tenant_id=current_user.tenant_id,
            part_id=transfer_request.part_id,
            location=transfer_request.to_location,
            quantity_on_hand=transfer_request.quantity,
            cost_per_unit=source_inventory.get("cost_per_unit", 0.0),
            total_value=transfer_request.quantity
            * source_inventory.get("cost_per_unit", 0.0),
        )
        await db.inventory.insert_one(new_inventory.dict())

    # Create movement record
    movement = InventoryMovement(
        tenant_id=current_user.tenant_id,
        part_id=transfer_request.part_id,
        movement_type="TRANSFER",
        from_location=transfer_request.from_location,
        to_location=transfer_request.to_location,
        quantity=transfer_request.quantity,
        unit_cost=source_inventory.get("cost_per_unit", 0.0),
        total_cost=transfer_request.quantity
        * source_inventory.get("cost_per_unit", 0.0),
        performed_by=current_user.id,
        notes=transfer_request.notes,
    )

    await db.inventory_movements.insert_one(movement.dict())

    await log_audit(
        current_user.id,
        AuditAction.UPDATE,
        "inventory_transfer",
        transfer_request.part_id,
        tenant_id=current_user.tenant_id,
    )

    return {"message": "Inventory transferred successfully", "movement_id": movement.id}


@api_router.get("/inventory/alerts", response_model=List[InventoryAlert])
async def get_inventory_alerts(current_user: User = Depends(get_current_user)):
    query = {"tenant_id": current_user.tenant_id}

    # Techs only see alerts relevant to them
    if current_user.role == UserRole.TECH:
        query["alert_type"] = {
            "$in": [AlertType.LOW_STOCK.value, AlertType.PART_REPLENISHMENT.value]
        }

    alerts = await db.inventory_alerts.find(query).sort("created_at", -1).to_list(100)
    return [InventoryAlert(**alert) for alert in alerts]


@api_router.post("/inventory/alerts/check")
async def check_inventory_alerts(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    alerts = await inventory_service.check_low_stock_alerts(current_user.tenant_id)
    return {"alerts_generated": len(alerts), "alerts": alerts}


@api_router.put("/inventory/alerts/{alert_id}/acknowledge")
async def acknowledge_alert(
    alert_id: str, current_user: User = Depends(get_current_user)
):
    result = await db.inventory_alerts.update_one(
        {"id": alert_id, "tenant_id": current_user.tenant_id},
        {
            "$set": {
                "is_acknowledged": True,
                "acknowledged_by": current_user.id,
                "acknowledged_at": datetime.now(timezone.utc),
            }
        },
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alert not found")

    return {"message": "Alert acknowledged successfully"}


# RMA Management Endpoints
@api_router.post("/rma", response_model=RMA)
async def create_rma(rma_data: dict, current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Generate RMA number
    year = datetime.now().year
    count = await db.rmas.count_documents({"tenant_id": current_user.tenant_id}) + 1
    rma_number = f"RMA-{year}-{str(count).zfill(4)}"

    rma_data["tenant_id"] = current_user.tenant_id
    rma_data["rma_number"] = rma_number
    rma_data["created_by"] = current_user.id

    rma = RMA(**rma_data)
    await db.rmas.insert_one(rma.dict())

    # Create notification for warehouse team
    await send_notification(
        current_user.tenant_id,  # This will notify all warehouse users
        NotificationType.RMA_UPDATE,
        "New RMA Created",
        f"RMA {rma_number} has been created and requires processing",
        {"rma_id": rma.id},
        current_user.tenant_id,
    )

    await log_audit(
        current_user.id,
        AuditAction.CREATE,
        "rma",
        rma.id,
        tenant_id=current_user.tenant_id,
    )

    return rma


@api_router.get("/rma", response_model=List[RMA])
async def get_rmas(
    status: Optional[RMAStatus] = None, current_user: User = Depends(get_current_user)
):
    query = {"tenant_id": current_user.tenant_id}

    if current_user.role == UserRole.TECH:
        # Techs can only see their own RMAs
        query["tech_id"] = current_user.id

    if status:
        query["status"] = status.value

    rmas = await db.rmas.find(query).sort("created_at", -1).to_list(100)
    return [RMA(**rma) for rma in rmas]


@api_router.put("/rma/{rma_id}/status")
async def update_rma_status(
    rma_id: str,
    status_update: RMAStatusUpdate,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        updated_rma = await inventory_service.process_rma_workflow(
            rma_id,
            status_update.new_status,
            status_update.dict(exclude={"rma_id", "new_status"}, exclude_none=True),
        )

        await log_audit(
            current_user.id,
            AuditAction.UPDATE,
            "rma",
            rma_id,
            tenant_id=current_user.tenant_id,
        )

        return updated_rma
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# Vendor Management Endpoints
@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(
    vendor_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    vendor_data["tenant_id"] = current_user.tenant_id
    vendor = Vendor(**vendor_data)
    await db.vendors.insert_one(vendor.dict())

    return vendor


@api_router.get("/vendors", response_model=List[Vendor])
async def get_vendors(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    vendors = await db.vendors.find(query).to_list(100)
    return [Vendor(**vendor) for vendor in vendors]


@api_router.post("/vendor-pickups", response_model=VendorPickup)
async def create_vendor_pickup(
    pickup_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Generate pickup number
    count = (
        await db.vendor_pickups.count_documents({"tenant_id": current_user.tenant_id})
        + 1
    )
    pickup_number = f"VP-{datetime.now().strftime('%Y%m')}-{str(count).zfill(3)}"

    pickup_data["tenant_id"] = current_user.tenant_id
    pickup_data["pickup_number"] = pickup_number

    pickup = VendorPickup(**pickup_data)
    await db.vendor_pickups.insert_one(pickup.dict())

    return pickup


@api_router.put("/vendor-pickups/{pickup_id}/status")
async def update_vendor_pickup_status(
    pickup_id: str, status_update: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        updated_pickup = await inventory_service.track_vendor_pickup(
            pickup_id, status_update
        )

        await log_audit(
            current_user.id,
            AuditAction.UPDATE,
            "vendor_pickup",
            pickup_id,
            tenant_id=current_user.tenant_id,
        )

        return updated_pickup
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@api_router.get("/vendor-pickups", response_model=List[VendorPickup])
async def get_vendor_pickups(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    pickups = (
        await db.vendor_pickups.find(query).sort("scheduled_date", -1).to_list(100)
    )
    return [VendorPickup(**pickup) for pickup in pickups]


# Enhanced Parts Management
@api_router.post("/enhanced-parts", response_model=EnhancedPart)
async def create_enhanced_part(
    part_data: dict, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.WAREHOUSE):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    part_data["tenant_id"] = current_user.tenant_id
    part = EnhancedPart(**part_data)
    await db.enhanced_parts.insert_one(part.dict())

    return part


@api_router.get("/enhanced-parts", response_model=List[EnhancedPart])
async def get_enhanced_parts(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    parts = await db.enhanced_parts.find(query).to_list(1000)
    return [EnhancedPart(**part) for part in parts]


# Enhanced Inventory Operations with auto-deduction
@api_router.post("/jobs/{job_id}/complete-with-parts")
async def complete_job_with_parts_deduction(
    job_id: str, completion_data: dict, current_user: User = Depends(get_current_user)
):
    if current_user.role != UserRole.TECH:
        raise HTTPException(status_code=403, detail="Only techs can complete jobs")

    # Verify job ownership
    job = await db.jobs.find_one({"id": job_id, "assigned_tech_id": current_user.id})
    if not job:
        raise HTTPException(
            status_code=404, detail="Job not found or not assigned to you"
        )

    parts_used = completion_data.get("parts_used", [])

    # Auto-deduct parts from inventory
    deduction_result = await inventory_service.auto_deduct_parts(
        job_id, parts_used, current_user.id
    )

    # Update job status and add parts used
    await db.jobs.update_one(
        {"id": job_id},
        {
            "$set": {
                "status": JobStatus.COMPLETED.value,
                "completed_date": datetime.now(timezone.utc),
                "parts_used": parts_used,
                "tech_notes": completion_data.get("tech_notes", ""),
                "updated_at": datetime.now(timezone.utc),
            }
        },
    )

    await log_audit(
        current_user.id,
        AuditAction.UPDATE,
        "job_complete",
        job_id,
        tenant_id=current_user.tenant_id,
    )

    return {
        "message": "Job completed successfully",
        "parts_deduction": deduction_result,
    }


@api_router.get("/inventory/movements", response_model=List[InventoryMovement])
async def get_inventory_movements(
    part_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
):
    query = {"tenant_id": current_user.tenant_id}

    if part_id:
        query["part_id"] = part_id
    if movement_type:
        query["movement_type"] = movement_type

    movements = (
        await db.inventory_movements.find(query)
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )
    return [InventoryMovement(**movement) for movement in movements]


# Asset Tag Scanning
@api_router.get("/assets/scan/{tag_code}")
async def scan_asset_tag(tag_code: str, current_user: User = Depends(get_current_user)):
    query = {
        "$or": [
            {"asset_tag.coam_id": tag_code},
            {"asset_tag.qr_code": tag_code},
            {"asset_tag.barcode": tag_code},
        ]
    }

    if current_user.role != UserRole.SUPER_ADMIN:
        query["tenant_id"] = current_user.tenant_id

    asset_data = await db.assets.find_one(query)
    if not asset_data:
        raise HTTPException(status_code=404, detail="Asset not found")

    return EnhancedAsset(**asset_data)


# Financial Management Endpoints
@api_router.post("/financial/revenue-collection")
async def collect_revenue(
    revenue_data: RevenueCollectionRequest,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        result = await financial_service.process_revenue_collection(
            current_user.tenant_id, revenue_data, current_user.id
        )

        await log_audit(
            current_user.id,
            AuditAction.CREATE,
            "revenue_collection",
            result["revenue_entry_id"],
            tenant_id=current_user.tenant_id,
        )

        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/financial/expense-submission")
async def submit_expense(
    amount: float = Form(...),
    category: str = Form(...),
    description: str = Form(...),
    machine_id: Optional[str] = Form(None),
    location_id: Optional[str] = Form(None),
    job_id: Optional[str] = Form(None),
    invoice_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Create ExpenseSubmissionRequest from form data
    try:
        expense_category = ExpenseCategory(category)
    except ValueError:
        raise HTTPException(
            status_code=400, detail=f"Invalid expense category: {category}"
        )

    expense_data = ExpenseSubmissionRequest(
        amount=amount,
        category=expense_category,
        description=description,
        machine_id=machine_id,
        location_id=location_id,
        job_id=job_id,
    )

    # Handle file upload if provided
    invoice_file_path = None
    if invoice_file:
        # Validate file type
        allowed_types = ["image/jpeg", "image/png", "image/jpg", "application/pdf"]
        if invoice_file.content_type not in allowed_types:
            raise HTTPException(
                status_code=400, detail="File type not supported for OCR"
            )

        # Save file
        file_id = str(uuid.uuid4())
        file_extension = (
            invoice_file.filename.split(".")[-1]
            if "." in invoice_file.filename
            else "jpg"
        )
        invoice_file_path = UPLOAD_DIR / f"invoices" / f"{file_id}.{file_extension}"
        invoice_file_path.parent.mkdir(exist_ok=True)

        with open(invoice_file_path, "wb") as buffer:
            shutil.copyfileobj(invoice_file.file, buffer)

        invoice_file_path = str(invoice_file_path)

    try:
        result = await financial_service.process_expense_with_ocr(
            current_user.tenant_id, expense_data, current_user.id, invoice_file_path
        )

        await log_audit(
            current_user.id,
            AuditAction.CREATE,
            "expense_submission",
            result["expense_entry_id"],
            tenant_id=current_user.tenant_id,
        )

        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/dashboard")
async def get_financial_dashboard(
    period_start: date, period_end: date, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        summary = await financial_service.get_financial_summary(
            current_user.tenant_id, period_start, period_end
        )
        return summary
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/asset-performance")
async def get_asset_performance_analytics(
    period_start: date, period_end: date, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        performance_data = await financial_service.calculate_asset_performance(
            current_user.tenant_id, period_start, period_end
        )
        return performance_data
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/financial/generate-report")
async def generate_financial_report(
    report_request: FinancialReportRequest,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Generate P&L statement
        pnl = await financial_service.generate_profit_loss_statement(
            current_user.tenant_id,
            report_request.period_start,
            report_request.period_end,
            current_user.id,
        )

        # For now, return the P&L data. Report generation (Excel/PDF) will be handled by frontend
        return pnl
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/predictive-insights")
async def get_predictive_insights(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        insights = await financial_service.get_predictive_insights(
            current_user.tenant_id
        )
        return insights
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/transactions")
async def get_financial_transactions(
    transaction_type: Optional[TransactionType] = None,
    machine_id: Optional[str] = None,
    location_id: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    query = {"tenant_id": current_user.tenant_id}

    if transaction_type:
        query["transaction_type"] = transaction_type.value
    if machine_id:
        query["machine_id"] = machine_id
    if location_id:
        query["location_id"] = location_id
    if start_date:
        query["created_at"] = {
            "$gte": datetime.combine(
                start_date, datetime.min.time().replace(tzinfo=timezone.utc)
            )
        }
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.combine(
                end_date, datetime.max.time().replace(tzinfo=timezone.utc)
            )
        else:
            query["created_at"] = {
                "$lte": datetime.combine(
                    end_date, datetime.max.time().replace(tzinfo=timezone.utc)
                )
            }

    transactions = (
        await db.financial_transactions.find(query)
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )
    return [FinancialTransaction(**transaction) for transaction in transactions]


@api_router.get("/financial/expenses")
async def get_expense_entries(
    category: Optional[ExpenseCategory] = None,
    requires_approval: Optional[bool] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    query = {"tenant_id": current_user.tenant_id}

    if category:
        query["category"] = category.value
    if requires_approval is not None:
        query["requires_approval"] = requires_approval
    if start_date:
        query["created_at"] = {
            "$gte": datetime.combine(
                start_date, datetime.min.time().replace(tzinfo=timezone.utc)
            )
        }
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.combine(
                end_date, datetime.max.time().replace(tzinfo=timezone.utc)
            )
        else:
            query["created_at"] = {
                "$lte": datetime.combine(
                    end_date, datetime.max.time().replace(tzinfo=timezone.utc)
                )
            }

    expenses = (
        await db.expense_entries.find(query)
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )
    return [ExpenseEntry(**expense) for expense in expenses]


@api_router.put("/financial/expenses/{expense_id}/approve")
async def approve_expense(
    expense_id: str, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    result = await db.expense_entries.update_one(
        {
            "id": expense_id,
            "tenant_id": current_user.tenant_id,
            "requires_approval": True,
        },
        {
            "$set": {
                "is_approved": True,
                "approved_by": current_user.id,
                "approved_at": datetime.now(timezone.utc),
            }
        },
    )

    if result.matched_count == 0:
        raise HTTPException(
            status_code=404, detail="Expense not found or already approved"
        )

    await log_audit(
        current_user.id,
        AuditAction.UPDATE,
        "expense_approval",
        expense_id,
        tenant_id=current_user.tenant_id,
    )

    return {"message": "Expense approved successfully"}


@api_router.post("/financial/ocr-process")
async def process_invoice_ocr(
    file: UploadFile = File(...), current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/jpg"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File type not supported for OCR")

    # Save file temporarily
    file_id = str(uuid.uuid4())
    file_extension = file.filename.split(".")[-1]
    temp_file_path = UPLOAD_DIR / f"temp_ocr" / f"{file_id}.{file_extension}"
    temp_file_path.parent.mkdir(exist_ok=True)

    with open(temp_file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        ocr_result = await financial_service.process_invoice_ocr(
            current_user.tenant_id, str(temp_file_path), current_user.id
        )
        return ocr_result
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"OCR processing failed: {str(e)}")
    finally:
        # Clean up temp file
        if temp_file_path.exists():
            temp_file_path.unlink()


# Financial report generation endpoints
@api_router.get("/financial/reports/profit-loss")
async def generate_profit_loss_report(
    period_start: date,
    period_end: date,
    format: str = "excel",  # excel or pdf
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Generate P&L statement
        pnl = await financial_service.generate_profit_loss_statement(
            current_user.tenant_id, period_start, period_end, current_user.id
        )

        filename = f"profit_loss_report_{period_start}_{period_end}"

        if format.lower() == "pdf":
            return await generate_pnl_pdf(pnl, filename)
        else:
            return await generate_pnl_excel(pnl, filename)

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/reports/expense-report")
async def generate_expense_report(
    period_start: date,
    period_end: date,
    format: str = "excel",
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get expenses from database
        query = {
            "tenant_id": current_user.tenant_id,
            "created_at": {
                "$gte": datetime.combine(
                    period_start, datetime.min.time().replace(tzinfo=timezone.utc)
                ),
                "$lte": datetime.combine(
                    period_end, datetime.max.time().replace(tzinfo=timezone.utc)
                ),
            },
        }

        if category:
            query["category"] = category

        expenses = (
            await db.expense_entries.find(query).sort("created_at", -1).to_list(None)
        )

        filename = f"expense_report_{period_start}_{period_end}"

        if format.lower() == "pdf":
            return await generate_expense_pdf(
                expenses, filename, period_start, period_end
            )
        else:
            return await generate_expense_excel(
                expenses, filename, period_start, period_end
            )

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/reports/revenue-analysis")
async def generate_revenue_analysis_report(
    period_start: date,
    period_end: date,
    format: str = "excel",
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get revenue data
        query = {
            "tenant_id": current_user.tenant_id,
            "transaction_type": "revenue",
            "created_at": {
                "$gte": datetime.combine(
                    period_start, datetime.min.time().replace(tzinfo=timezone.utc)
                ),
                "$lte": datetime.combine(
                    period_end, datetime.max.time().replace(tzinfo=timezone.utc)
                ),
            },
        }

        revenue_data = (
            await db.financial_transactions.find(query)
            .sort("created_at", -1)
            .to_list(None)
        )

        # Get asset performance data
        asset_performance = await financial_service.calculate_asset_performance(
            current_user.tenant_id, period_start, period_end
        )

        filename = f"revenue_analysis_{period_start}_{period_end}"

        if format.lower() == "pdf":
            return await generate_revenue_analysis_pdf(
                revenue_data, asset_performance, filename, period_start, period_end
            )
        else:
            return await generate_revenue_analysis_excel(
                revenue_data, asset_performance, filename, period_start, period_end
            )

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/financial/reports/commission-report")
async def generate_commission_report(
    period_start: date,
    period_end: date,
    format: str = "excel",
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get commission transactions
        query = {
            "tenant_id": current_user.tenant_id,
            "transaction_type": {"$in": ["revenue", "commission"]},
            "created_at": {
                "$gte": datetime.combine(
                    period_start, datetime.min.time().replace(tzinfo=timezone.utc)
                ),
                "$lte": datetime.combine(
                    period_end, datetime.max.time().replace(tzinfo=timezone.utc)
                ),
            },
        }

        transactions = (
            await db.financial_transactions.find(query)
            .sort("created_at", -1)
            .to_list(None)
        )

        filename = f"commission_report_{period_start}_{period_end}"

        if format.lower() == "pdf":
            return await generate_commission_pdf(
                transactions, filename, period_start, period_end
            )
        else:
            return await generate_commission_excel(
                transactions, filename, period_start, period_end
            )

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Helper functions for report generation
async def generate_pnl_excel(pnl_data, filename):
    buffer = io.BytesIO()

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss Statement"

    # Headers
    ws["A1"] = "Profit & Loss Statement"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = f"Period: {pnl_data.period_start} to {pnl_data.period_end}"

    # Revenue section
    row = 5
    ws[f"A{row}"] = "REVENUE"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    # Revenue by source
    for source, amount in pnl_data.revenue_by_source.items():
        ws[f"A{row}"] = f"Revenue from {source}"
        ws[f"B{row}"] = amount
        row += 1

    ws[f"A{row}"] = "Total Revenue"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = pnl_data.gross_revenue
    ws[f"B{row}"].font = Font(bold=True)
    row += 2

    # Expenses section
    ws[f"A{row}"] = "EXPENSES"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    # Expenses by category
    for category, amount in pnl_data.expenses_by_category.items():
        ws[f"A{row}"] = f"{category.title()} expenses"
        ws[f"B{row}"] = amount
        row += 1

    ws[f"A{row}"] = "Total Expenses"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = pnl_data.total_expenses
    ws[f"B{row}"].font = Font(bold=True)
    row += 2

    # Net profit
    ws[f"A{row}"] = "NET PROFIT"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    ws[f"B{row}"] = pnl_data.net_profit
    ws[f"B{row}"].font = Font(bold=True, size=14)

    # Format currency columns
    for row_num in range(1, ws.max_row + 1):
        if ws[f"B{row_num}"].value and isinstance(
            ws[f"B{row_num}"].value, (int, float)
        ):
            ws[f"B{row_num}"].number_format = "$#,##0.00"

    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


async def generate_expense_excel(expenses, filename, period_start, period_end):
    buffer = io.BytesIO()

    # Create DataFrame from expenses
    expense_data = []
    for expense in expenses:
        expense_data.append(
            {
                "Date": (
                    expense.get("created_at", "").strftime("%Y-%m-%d")
                    if expense.get("created_at")
                    else ""
                ),
                "Category": expense.get("category", ""),
                "Amount": expense.get("amount", 0),
                "Description": expense.get("description", ""),
                "Machine ID": expense.get("machine_id", ""),
                "Location ID": expense.get("location_id", ""),
                "Status": "Approved" if expense.get("approved", False) else "Pending",
            }
        )

    df = pd.DataFrame(expense_data)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Write summary sheet
        summary_df = pd.DataFrame(
            [
                ["Report Period", f"{period_start} to {period_end}"],
                ["Total Expenses", len(expenses)],
                ["Total Amount", df["Amount"].sum() if not df.empty else 0],
            ],
            columns=["Metric", "Value"],
        )

        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write detailed data
        if not df.empty:
            df.to_excel(writer, sheet_name="Expense Details", index=False)

    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


async def generate_revenue_analysis_excel(
    revenue_data, asset_performance, filename, period_start, period_end
):
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Revenue summary
        if revenue_data:
            revenue_df = pd.DataFrame(
                [
                    {
                        "Date": (
                            item.get("created_at", "").strftime("%Y-%m-%d")
                            if item.get("created_at")
                            else ""
                        ),
                        "Machine ID": item.get("machine_id", ""),
                        "Location ID": item.get("location_id", ""),
                        "Amount": item.get("amount", 0),
                        "Collection Method": item.get("metadata", {}).get(
                            "collection_method", "Unknown"
                        ),
                    }
                    for item in revenue_data
                ]
            )
            revenue_df.to_excel(writer, sheet_name="Revenue Data", index=False)

        # Asset performance
        if asset_performance:
            performance_df = pd.DataFrame(
                [
                    {
                        "Asset Name": asset.asset_name,
                        "COAM ID": asset.coam_id,
                        "Location": asset.location_name,
                        "Total Revenue": asset.total_revenue,
                        "ROI %": asset.roi_percentage,
                        "Performance Rank": asset.performance_rank,
                        "Category": asset.performance_category,
                    }
                    for asset in asset_performance
                ]
            )
            performance_df.to_excel(writer, sheet_name="Asset Performance", index=False)

    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


async def generate_commission_excel(transactions, filename, period_start, period_end):
    buffer = io.BytesIO()

    # Process commission data
    commission_data = []
    total_revenue = 0
    total_ml_share = 0
    total_location_share = 0

    for transaction in transactions:
        if transaction.get("transaction_type") == "revenue":
            amount = transaction.get("amount", 0)
            total_revenue += amount

            # Calculate commission split (assuming 60/40 split)
            ml_share = amount * 0.6
            location_share = amount * 0.4

            total_ml_share += ml_share
            total_location_share += location_share

            commission_data.append(
                {
                    "Date": (
                        transaction.get("created_at", "").strftime("%Y-%m-%d")
                        if transaction.get("created_at")
                        else ""
                    ),
                    "Machine ID": transaction.get("machine_id", ""),
                    "Location ID": transaction.get("location_id", ""),
                    "Gross Revenue": amount,
                    "ML Share (60%)": ml_share,
                    "Location Share (40%)": location_share,
                }
            )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Summary sheet
        summary_df = pd.DataFrame(
            [
                ["Report Period", f"{period_start} to {period_end}"],
                ["Total Gross Revenue", total_revenue],
                ["Total ML Share", total_ml_share],
                ["Total Location Share", total_location_share],
                ["Number of Collections", len(commission_data)],
            ],
            columns=["Metric", "Value"],
        )

        summary_df.to_excel(writer, sheet_name="Commission Summary", index=False)

        # Detailed commission data
        if commission_data:
            commission_df = pd.DataFrame(commission_data)
            commission_df.to_excel(writer, sheet_name="Commission Details", index=False)

    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


# PDF generation functions (simplified for now, can be enhanced)
async def generate_pnl_pdf(pnl_data, filename):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title = Paragraph("Profit & Loss Statement", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    # Period
    period = Paragraph(
        f"Period: {pnl_data.period_start} to {pnl_data.period_end}", styles["Normal"]
    )
    story.append(period)
    story.append(Spacer(1, 12))

    # Revenue section
    revenue_title = Paragraph("REVENUE", styles["Heading2"])
    story.append(revenue_title)

    for source, amount in pnl_data.revenue_by_source.items():
        item_text = f"Revenue from {source}: ${amount:,.2f}"
        story.append(Paragraph(item_text, styles["Normal"]))

    total_revenue = Paragraph(
        f"<b>Total Revenue: ${pnl_data.gross_revenue:,.2f}</b>", styles["Normal"]
    )
    story.append(total_revenue)
    story.append(Spacer(1, 12))

    # Expenses section
    expense_title = Paragraph("EXPENSES", styles["Heading2"])
    story.append(expense_title)

    for category, amount in pnl_data.expenses_by_category.items():
        item_text = f"{category.title()} expenses: ${amount:,.2f}"
        story.append(Paragraph(item_text, styles["Normal"]))

    total_expenses = Paragraph(
        f"<b>Total Expenses: ${pnl_data.total_expenses:,.2f}</b>", styles["Normal"]
    )
    story.append(total_expenses)
    story.append(Spacer(1, 12))

    # Net profit
    net_profit = Paragraph(
        f"<b>NET PROFIT: ${pnl_data.net_profit:,.2f}</b>", styles["Heading1"]
    )
    story.append(net_profit)

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


async def generate_expense_pdf(expenses, filename, period_start, period_end):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Expense Report", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    period = Paragraph(f"Period: {period_start} to {period_end}", styles["Normal"])
    story.append(period)
    story.append(Spacer(1, 12))

    # Summary
    total_amount = sum(expense.get("amount", 0) for expense in expenses)
    summary = Paragraph(
        f"Total Expenses: ${total_amount:,.2f} ({len(expenses)} entries)",
        styles["Heading2"],
    )
    story.append(summary)
    story.append(Spacer(1, 12))

    # Table data
    data = [["Date", "Category", "Amount", "Description"]]
    for expense in expenses[:20]:  # Limit to first 20 for PDF
        date_str = (
            expense.get("created_at", "").strftime("%Y-%m-%d")
            if expense.get("created_at")
            else ""
        )
        data.append(
            [
                date_str,
                expense.get("category", ""),
                f"${expense.get('amount', 0):,.2f}",
                (
                    expense.get("description", "")[:50] + "..."
                    if len(expense.get("description", "")) > 50
                    else expense.get("description", "")
                ),
            ]
        )

    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    story.append(table)
    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


async def generate_revenue_analysis_pdf(
    revenue_data, asset_performance, filename, period_start, period_end
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Revenue Analysis Report", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    period = Paragraph(f"Period: {period_start} to {period_end}", styles["Normal"])
    story.append(period)
    story.append(Spacer(1, 12))

    # Summary
    total_revenue = sum(item.get("amount", 0) for item in revenue_data)
    summary = Paragraph(f"Total Revenue: ${total_revenue:,.2f}", styles["Heading2"])
    story.append(summary)
    story.append(Spacer(1, 12))

    # Asset performance summary
    if asset_performance:
        performance_title = Paragraph("Top Performing Assets", styles["Heading2"])
        story.append(performance_title)

        for asset in asset_performance[:5]:  # Top 5
            asset_text = f"{asset.asset_name}: ${asset.total_revenue:,.2f} (ROI: {asset.roi_percentage:.1f}%)"
            story.append(Paragraph(asset_text, styles["Normal"]))

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


async def generate_commission_pdf(transactions, filename, period_start, period_end):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("Commission Report", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    period = Paragraph(f"Period: {period_start} to {period_end}", styles["Normal"])
    story.append(period)
    story.append(Spacer(1, 12))

    # Calculate totals
    revenue_transactions = [
        t for t in transactions if t.get("transaction_type") == "revenue"
    ]
    total_revenue = sum(t.get("amount", 0) for t in revenue_transactions)
    total_ml_share = total_revenue * 0.6
    total_location_share = total_revenue * 0.4

    summary = Paragraph(
        f"""
    Total Revenue: ${total_revenue:,.2f}<br/>
    ML Share (60%): ${total_ml_share:,.2f}<br/>
    Location Share (40%): ${total_location_share:,.2f}
    """,
        styles["Normal"],
    )
    story.append(summary)

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
    )


# Notification and Automation Endpoints
@api_router.post("/notifications")
async def create_notification(
    request: CreateNotificationRequest, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        notification = await notification_service.create_notification(
            current_user.tenant_id, request
        )
        return notification
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/notifications")
async def get_user_notifications(
    limit: int = 50,
    unread_only: bool = False,
    current_user: User = Depends(get_current_user),
):
    try:
        notifications = await notification_service.get_user_notifications(
            current_user.id, current_user.tenant_id, limit, unread_only
        )
        return notifications
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str, current_user: User = Depends(get_current_user)
):
    try:
        success = await notification_service.mark_notification_read(
            current_user.id, notification_id
        )
        if success:
            return {"message": "Notification marked as read"}
        else:
            raise HTTPException(status_code=404, detail="Notification not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/notifications/stats")
async def get_notification_stats(current_user: User = Depends(get_current_user)):
    try:
        stats = await notification_service.get_notification_stats(
            current_user.id, current_user.tenant_id
        )
        return stats
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Location Notification Settings
@api_router.get("/locations/{location_id}/notification-settings")
async def get_location_notification_settings(
    location_id: str, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        settings = await notification_service.get_location_notification_settings(
            current_user.tenant_id, location_id
        )
        if settings:
            return settings
        else:
            raise HTTPException(status_code=404, detail="Settings not found")
    except HTTPException:
        raise  # Re-raise HTTPExceptions (like 404) without modification
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.put("/locations/{location_id}/notification-settings")
async def update_location_notification_settings(
    location_id: str,
    settings_request: UpdateNotificationSettingsRequest,
    location_name: str = Form(...),
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        settings = await notification_service.create_or_update_location_settings(
            current_user.tenant_id, location_id, location_name, settings_request
        )
        return settings
    except HTTPException:
        raise  # Re-raise HTTPExceptions without modification
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/locations/notification-settings")
async def get_all_location_settings(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        settings_list = await notification_service.get_all_location_settings(
            current_user.tenant_id
        )
        return settings_list
    except HTTPException:
        raise  # Re-raise HTTPExceptions without modification
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Renewal Management
@api_router.post("/renewals")
async def create_renewal_item(
    request: CreateRenewalItemRequest,
    location_id: str = Form(...),
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        renewal_item = await automation_service.create_renewal_item(
            current_user.tenant_id, location_id, request
        )
        return renewal_item
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/renewals")
async def get_renewal_items(
    location_id: Optional[str] = None, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        renewals = await automation_service.get_renewal_items(
            current_user.tenant_id, location_id
        )
        return renewals
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.put("/renewals/{renewal_id}/complete")
async def mark_renewal_completed(
    renewal_id: str, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        success = await automation_service.mark_renewal_completed(
            current_user.tenant_id, renewal_id
        )
        if success:
            return {"message": "Renewal marked as completed"}
        else:
            raise HTTPException(status_code=404, detail="Renewal not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Automation Management
@api_router.post("/automation/run-checks")
async def run_automation_checks(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        await automation_service.run_automated_checks(current_user.tenant_id)
        return {"message": "Automation checks completed"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/automation/stats")
async def get_automation_stats(current_user: User = Depends(get_current_user)):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        stats = await automation_service.get_automation_stats(current_user.tenant_id)
        return stats
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Job Repost Management
@api_router.get("/job-reposts")
async def get_job_reposts(
    requires_approval: Optional[bool] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        query = {"tenant_id": current_user.tenant_id}
        if requires_approval is not None:
            query["requires_approval"] = requires_approval

        reposts = (
            await db.job_repost_records.find(query)
            .sort("created_at", -1)
            .limit(limit)
            .to_list(limit)
        )
        return [JobRepostRecord(**repost) for repost in reposts]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.put("/job-reposts/{repost_id}/approve")
async def approve_job_repost(
    repost_id: str, current_user: User = Depends(get_current_user)
):
    if not check_permission(current_user, UserRole.DISPATCH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        result = await db.job_repost_records.update_one(
            {"id": repost_id, "tenant_id": current_user.tenant_id},
            {
                "$set": {
                    "approved_by": current_user.id,
                    "approved_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                }
            },
        )

        if result.modified_count > 0:
            return {"message": "Job repost approved"}
        else:
            raise HTTPException(status_code=404, detail="Repost record not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Mobile App Integration Webhooks
@api_router.post("/webhooks/mobile-notifications")
async def mobile_notification_webhook(payload: dict):
    """Webhook endpoint for mobile app notifications"""
    logger.info(f"Mobile notification webhook received: {payload}")

    # Process mobile notification delivery
    # This would integrate with actual mobile push notification service

    return {"status": "received", "message": "Mobile notification processed"}


@api_router.post("/webhooks/mobile-delivery-status")
async def mobile_delivery_status_webhook(payload: dict):
    """Webhook endpoint for mobile delivery status updates"""
    logger.info(f"Mobile delivery status webhook: {payload}")

    # Update delivery status in database
    if payload.get("notification_id"):
        await db.webhook_deliveries.update_one(
            {"notification_id": payload["notification_id"]},
            {
                "$set": {
                    "status": payload.get("status", "unknown"),
                    "response_code": payload.get("response_code"),
                    "last_attempt_at": datetime.now(timezone.utc),
                }
            },
        )

    return {"status": "processed"}


# Real-time notification endpoints for frontend
@api_router.get("/notifications/realtime")
async def get_realtime_notifications(current_user: User = Depends(get_current_user)):
    """Get real-time notifications for frontend polling"""
    try:
        # Get recent unread notifications
        recent_notifications = await notification_service.get_user_notifications(
            current_user.id, current_user.tenant_id, limit=10, unread_only=True
        )

        # Get notification stats
        stats = await notification_service.get_notification_stats(
            current_user.id, current_user.tenant_id
        )

        return {
            "recent_notifications": recent_notifications,
            "stats": stats,
            "timestamp": datetime.now(timezone.utc),
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# System alerts for urgent issues
@api_router.post("/alerts/urgent")
async def create_urgent_alert(
    title: str = Form(...),
    message: str = Form(...),
    location_id: Optional[str] = Form(None),
    machine_id: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
):
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Create urgent alert notification
        request = CreateNotificationRequest(
            type=NotificationType.URGENT_ISSUE,
            priority=NotificationPriority.URGENT,
            title=title,
            message=message,
            location_ids=[location_id] if location_id else [],
            role_targets=["TECH", "DISPATCH", "ML_ADMIN"],
            channels=[
                NotificationChannel.SMS,
                NotificationChannel.WEB_PUSH,
                NotificationChannel.MOBILE_PUSH,
            ],
            data={
                "machine_id": machine_id,
                "created_by": current_user.id,
                "created_by_name": f"{current_user.first_name} {current_user.last_name}",
            },
            related_entity_type="machine" if machine_id else "alert",
            related_entity_id=machine_id or str(uuid.uuid4()),
        )

        notification = await notification_service.create_notification(
            current_user.tenant_id, request
        )
        return notification
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# =============================================================================
# BILLING AND SUBSCRIPTION MANAGEMENT ENDPOINTS
# =============================================================================


# Billing Dashboard
@api_router.get("/billing/dashboard", response_model=BillingDashboardResponse)
async def get_billing_dashboard(current_user: User = Depends(get_current_user)):
    """Get billing dashboard for current user's tenant"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get or create customer for this user's tenant
        customer = await billing_service.get_customer_by_email(
            current_user.tenant_id, current_user.email
        )
        if not customer:
            # Create customer if doesn't exist
            create_request = CreateCustomerRequest(
                email=current_user.email,
                name=f"{current_user.first_name} {current_user.last_name}",
            )
            customer = await billing_service.create_customer(
                current_user.tenant_id, create_request
            )

        dashboard = await billing_service.get_billing_dashboard(
            current_user.tenant_id, customer.id
        )
        return dashboard if dashboard else BillingDashboardResponse(customer=customer)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Pricing Plans
@api_router.get("/billing/pricing-plans", response_model=List[PricingPlan])
async def get_pricing_plans():
    """Get all available pricing plans"""
    try:
        return await billing_service.get_pricing_plans()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/billing/packages")
async def get_subscription_packages():
    """Get predefined subscription packages"""
    return stripe_integration.get_subscription_packages()


# Subscription Management
@api_router.post("/billing/subscriptions", response_model=Subscription)
async def create_subscription(
    request: CreateSubscriptionRequest, current_user: User = Depends(get_current_user)
):
    """Create new subscription"""
    if not check_permission(current_user, UserRole.ML_ADMIN):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        return await billing_service.create_subscription(
            current_user.tenant_id, request
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/billing/subscriptions/{subscription_id}", response_model=Subscription)
async def get_subscription(
    subscription_id: str, current_user: User = Depends(get_current_user)
):
    """Get subscription details"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    subscription = await billing_service.get_subscription(
        current_user.tenant_id, subscription_id
    )
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")
    return subscription


@api_router.put("/billing/subscriptions/{subscription_id}", response_model=Subscription)
async def update_subscription(
    subscription_id: str,
    request: UpdateSubscriptionRequest,
    current_user: User = Depends(get_current_user),
):
    """Update subscription details"""
    if not check_permission(current_user, UserRole.ML_ADMIN):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        subscription = await billing_service.update_subscription(
            current_user.tenant_id, subscription_id, request
        )
        if not subscription:
            raise HTTPException(status_code=404, detail="Subscription not found")
        return subscription
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.delete("/billing/subscriptions/{subscription_id}")
async def cancel_subscription(
    subscription_id: str, current_user: User = Depends(get_current_user)
):
    """Cancel subscription"""
    if not check_permission(current_user, UserRole.ML_ADMIN):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        await billing_service.cancel_subscription(
            current_user.tenant_id, subscription_id
        )
        return {"message": "Subscription canceled successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Stripe Checkout Integration
@api_router.post("/billing/checkout/subscription")
async def create_subscription_checkout(
    subscription_id: str = Form(...),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    """Create Stripe checkout session for subscription payment"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get host URL from request
        host_url = str(request.base_url).rstrip("/")

        # Create success and cancel URLs
        success_url = f"{host_url}/billing/success?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{host_url}/billing/cancel"

        session = await stripe_integration.create_subscription_checkout_session(
            tenant_id=current_user.tenant_id,
            subscription_id=subscription_id,
            host_url=host_url,
            success_url=success_url,
            cancel_url=cancel_url,
        )

        return {"url": session.url, "session_id": session.session_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/billing/checkout/status/{session_id}")
async def get_checkout_status(
    session_id: str, request: Request, current_user: User = Depends(get_current_user)
):
    """Get checkout session status"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        host_url = str(request.base_url).rstrip("/")
        status = await stripe_integration.get_checkout_status(session_id, host_url)
        return status
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Invoice Management
@api_router.get("/billing/invoices", response_model=List[Invoice])
async def get_customer_invoices(
    limit: int = 10, current_user: User = Depends(get_current_user)
):
    """Get customer invoices"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get customer
        customer = await billing_service.get_customer_by_email(
            current_user.tenant_id, current_user.email
        )
        if not customer:
            return []

        return await billing_service.get_customer_invoices(
            current_user.tenant_id, customer.id, limit
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/billing/invoices/{invoice_id}/pay")
async def create_invoice_payment(
    invoice_id: str,
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    """Create payment for invoice"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Get host URL from request
        host_url = str(request.base_url).rstrip("/")

        # Create success and cancel URLs
        success_url = f"{host_url}/billing/success?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{host_url}/billing/cancel"

        session = await stripe_integration.create_invoice_checkout_session(
            tenant_id=current_user.tenant_id,
            invoice_id=invoice_id,
            host_url=host_url,
            success_url=success_url,
            cancel_url=cancel_url,
        )

        return {"url": session.url, "session_id": session.session_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Admin Billing Statistics (Super Admin only)
@api_router.get("/billing/admin/stats", response_model=BillingStatsResponse)
async def get_billing_stats(
    tenant_id: Optional[str] = None, current_user: User = Depends(get_current_user)
):
    """Get billing statistics for admin dashboard"""
    if not check_permission(current_user, UserRole.SUPER_ADMIN):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        # Super admin can view all tenants, others only their own
        stats_tenant_id = (
            tenant_id
            if current_user.role == UserRole.SUPER_ADMIN
            else current_user.tenant_id
        )
        return await billing_service.get_billing_stats(stats_tenant_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Usage Tracking
@api_router.post("/billing/usage/record")
async def record_usage(
    subscription_id: str = Form(...),
    machine_count: Optional[int] = Form(None),
    user_count: Optional[int] = Form(None),
    api_calls: int = Form(0),
    reports_generated: int = Form(0),
    notifications_sent: int = Form(0),
    current_user: User = Depends(get_current_user),
):
    """Record usage for billing"""
    if not check_permission(current_user, UserRole.TECH):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    try:
        usage = await billing_service.record_usage(
            tenant_id=current_user.tenant_id,
            subscription_id=subscription_id,
            machine_count=machine_count,
            user_count=user_count,
            api_calls=api_calls,
            reports_generated=reports_generated,
            notifications_sent=notifications_sent,
        )
        return {"message": "Usage recorded successfully", "usage_id": usage.id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Stripe Webhook Endpoint
@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhooks"""
    try:
        result = await stripe_integration.handle_webhook(request)
        return result
    except Exception as e:
        logger.error(f"Stripe webhook error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# Health check
@api_router.get("/health")
async def health_check():
    try:
        # Test database connection
        await db.command("ping")
        return {"status": "healthy", "timestamp": datetime.now(timezone.utc)}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}


# Include the router in the main app
app.include_router(api_router)

# Add middlewares
app.add_middleware(TenantMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup_event():
    global inventory_service, financial_service, notification_service, automation_service, billing_service, stripe_integration
    inventory_service = InventoryService(db)
    financial_service = FinancialService(db)
    notification_service = NotificationService(db)
    automation_service = AutomationService(db, notification_service)
    billing_service = BillingService(db)
    stripe_integration = StripeIntegrationService(billing_service)

    # Initialize default pricing plans
    await billing_service.initialize_default_plans()

    # Create indexes for better performance
    await db.users.create_index([("email", 1)], unique=True)
    await db.users.create_index([("tenant_id", 1)])
    await db.jobs.create_index([("tenant_id", 1), ("status", 1)])
    await db.jobs.create_index([("assigned_tech_id", 1), ("status", 1)])
    await db.machines.create_index(
        [("tenant_id", 1), ("serial_number", 1)], unique=True
    )
    await db.machines.create_index([("qr_code", 1)])
    await db.machines.create_index([("barcode", 1)])
    await db.locations.create_index([("tenant_id", 1)])
    await db.parts.create_index([("tenant_id", 1), ("part_number", 1)])
    await db.inventory.create_index([("tenant_id", 1), ("part_id", 1)])
    await db.truck_inventory.create_index([("tech_id", 1), ("part_id", 1)])
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("timestamp", -1)])

    # Enhanced inventory indexes
    await db.assets.create_index([("tenant_id", 1), ("asset_type", 1)])
    await db.assets.create_index([("asset_tag.coam_id", 1)], unique=True)
    await db.assets.create_index([("asset_tag.qr_code", 1)])
    await db.assets.create_index([("asset_tag.barcode", 1)])
    await db.inventory_alerts.create_index([("tenant_id", 1), ("is_acknowledged", 1)])
    await db.inventory_movements.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.rmas.create_index([("tenant_id", 1), ("status", 1)])
    await db.vendor_pickups.create_index([("tenant_id", 1), ("scheduled_date", 1)])
    await db.asset_lifecycle.create_index([("asset_id", 1), ("event_date", -1)])

    # Billing system indexes
    await db.customers.create_index([("tenant_id", 1), ("email", 1)], unique=True)
    await db.customers.create_index([("stripe_customer_id", 1)])
    await db.subscriptions.create_index([("tenant_id", 1), ("customer_id", 1)])
    await db.subscriptions.create_index([("tenant_id", 1), ("status", 1)])
    await db.subscriptions.create_index([("stripe_subscription_id", 1)])
    await db.invoices.create_index([("tenant_id", 1), ("customer_id", 1)])
    await db.invoices.create_index([("tenant_id", 1), ("status", 1)])
    await db.invoices.create_index([("stripe_invoice_id", 1)])
    await db.payment_transactions.create_index([("tenant_id", 1), ("customer_id", 1)])
    await db.payment_transactions.create_index([("stripe_session_id", 1)])
    await db.payment_transactions.create_index([("status", 1), ("created_at", -1)])
    await db.usage_records.create_index([("tenant_id", 1), ("subscription_id", 1)])
    await db.pricing_plans.create_index([("tier", 1), ("is_active", 1)])

    # Financial system indexes
    await db.financial_transactions.create_index(
        [("tenant_id", 1), ("transaction_type", 1), ("created_at", -1)]
    )
    await db.financial_transactions.create_index(
        [("machine_id", 1), ("transaction_type", 1)]
    )
    await db.financial_transactions.create_index(
        [("location_id", 1), ("transaction_type", 1)]
    )
    await db.revenue_entries.create_index([("tenant_id", 1), ("collection_date", -1)])
    await db.expense_entries.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.expense_entries.create_index([("category", 1), ("tenant_id", 1)])
    await db.invoice_ocr.create_index([("tenant_id", 1), ("processed_at", -1)])
    await db.profit_loss_statements.create_index(
        [("tenant_id", 1), ("period_start", -1)]
    )

    logger.info("Application startup completed")

    # Create default super admin if not exists
    super_admin = await db.users.find_one({"role": UserRole.SUPER_ADMIN})
    if not super_admin:
        admin_user = User(
            email="admin@coamsaas.com",
            first_name="Super",
            last_name="Admin",
            role=UserRole.SUPER_ADMIN,
            tenant_id=None,
        )
        admin_dict = admin_user.dict()
        admin_dict["password_hash"] = hash_password("admin123")  # Change in production
        await db.users.insert_one(admin_dict)
        logger.info("Default super admin created")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
